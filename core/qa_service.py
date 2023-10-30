
from openpyxl import load_workbook
import difflib
from utils import config_util as cfg

def question(query_type,text):
    qa = QAService()
    answer = qa.question(query_type,text)
    return answer

class QAService:
    
    def __init__(self):
         # 人设提问关键字
        self.attribute_keyword = [
            [['你叫什么名字', '你的名字是什么'], 'name'],
            [['你是男的还是女的', '你是男生还是女生', '你的性别是什么', '你是男生吗', '你是女生吗', '你是男的吗', '你是女的吗', '你是男孩子吗', '你是女孩子吗', ], 'gender', ],
            [['你今年多大了', '你多大了', '你今年多少岁', '你几岁了', '你今年几岁了', '你今年几岁了', '你什么时候出生', '你的生日是什么', '你的年龄'], 'age', ],
            [['你的家乡在哪', '你的家乡是什么', '你家在哪', '你住在哪', '你出生在哪', '你的出生地在哪', '你的出生地是什么', ], 'birth', ],
            [['你的生肖是什么', '你属什么', ], 'zodiac', ],
            [['你是什么座', '你是什么星座', '你的星座是什么', ], 'constellation', ],
            [['你是做什么的', '你的职业是什么', '你是干什么的', '你的职位是什么', '你的工作是什么', '你是做什么工作的'], 'job', ],
            [['你的爱好是什么', '你有爱好吗', '你喜欢什么', '你喜欢做什么'], 'hobby'],
            [['联系方式', '联系你们', '怎么联系客服', '有没有客服'], 'contact']
        ]

        self.command_keyword = [
            [['关闭', '再见', '你走吧'], 'stop'],
            [['静音', '闭嘴', '我想静静'], 'mute'],
            [['取消静音', '你在哪呢', '你可以说话了'], 'unmute'],
            [['换个性别', '换个声音'], 'changeVoice']
        ]

        # 商品提问关键字
        self.explain_keyword = [
            [['是什么'], 'intro'],
            [['怎么用', '使用场景', '有什么作用'], 'usage'],
            [['怎么卖', '多少钱', '售价'], 'price'],
            [['便宜点', '优惠', '折扣', '促销'], 'discount'],
            [['质量', '保证', '担保'], 'promise'],
            [['特点', '优点'], 'character'],
        ]

    def question(self, query_type, text):
        answer = None
        if query_type == 'qa':
            answer_dict = self.__read_qna(cfg.config['interact']['QnA'])
            answer = self.__get_keyword(answer_dict, text)
        elif query_type == 'Persona':
            answer_dict = self.attribute_keyword
            answer = self.__get_keyword(answer_dict, text)
        elif query_type == 'command':
            answer = self.__get_keyword(self.command_keyword, text)
        elif query_type == 'goods':
            items = self.__get_item_list()
            if len(items) > 0:
                item = items[0]

                # 跨商品物品问答匹配
                for ite in items:
                    name = ite["name"]
                    if name != item["name"]:
                        if name in text or self.__string_similar(text, name) > 0.6:
                            item = ite
                            break

                # 商品介绍问答
                keyword = self.__get_keyword(self.explain_keyword, text)
                if keyword is not None:
                    try:
                        return item["explain"][keyword]
                    except BaseException as e:
                        print(e)

                # 商品问答
                answer = self.__get_keyword(self.__read_qna(item["QnA"]), text)
                if answer is not None:
                    return answer


        return answer
    
    def __get_item_list(self) -> list:
        items = []
        for item in cfg.config["items"]:
            if item["enabled"]:
                items.append(item)
        return items

    def __read_qna(self, filename):
        qna = []
        try:
            wb = load_workbook(filename)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= 2:
                    qna.append([row[0].split(";"), row[1]])
        except BaseException as e:
            print(f"无法读取Q&A文件 {filename} -> {e}")
        return qna

    def __get_keyword(self, keyword_dict, text):
        last_similar = 0
        last_answer = ''
        for qa in keyword_dict:
            for quest in qa[0]:
                similar = self.__string_similar(text, quest)
                if quest in text:
                    similar += 0.3
                if similar > last_similar:
                    last_similar = similar
                    last_answer = qa[1]
        if last_similar >= 0.6:
            return last_answer
        return None


    def __string_similar(self, s1, s2):
        return difflib.SequenceMatcher(None, s1, s2).quick_ratio()


